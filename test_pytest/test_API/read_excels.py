
import openpyxl

#
# def readexecl(filename, sheetname):
#     wb = openpyxl.load_workbook(filename)
#     # 指定要读取内容的sheet
#     sheetobj = wb[sheetname]
#     # 读取所有的内容
#     # 获取最大行
#     rows = sheetobj.max_row
#     # 获取最大列
#     columns = sheetobj.max_column
#     list = []
#     # 使用for循环，
#     for row in range(2, rows + 1):
#         # 定义一个列表，用来存放我们获取的每一个数据
#         new_list = []
#         # 列的处理
#         for column in range(1, columns + 1):
#             # 取值
#             new_list.append(sheetobj.cell(row, column).value)
#         list.append(new_list)
#
#     return list
#
# import xlrd, os
# '''
# 获取excel用例数据
#
# '''
#
#
# def get_case_data():
#     #  拼接读取文件的绝对路径
#     case_path = os.path.join(os.path.dirname(__file__), r'D:\pycharm-workspace\test_pytest\test_case\cases.xlsx')
#     #  读取文件
#     book = xlrd.open_workbook(case_path)
#     #  指定读取文件中的页名
#     sheet = book.sheet_by_name('adminLogin')
#     case = []
#     for i in range(0, sheet.nrows):
#         # 判断过滤掉标注的名称，是名称过滤
#         if sheet.row_values(i)[0] == 'URL':
#             pass
#         else:
#             #  如果不是记录出来使用数据
#             case.append(sheet.row_values(i))
#     return case
#
#
# # 调用获取测试用例数据
# case_data = get_case_data()
# print(case_data)
#

import pytest
import xlrd

def get_data():
    filename = r'D:\pycharm-workspace\test_pytest\test_case\cases.xlsx'
    wb = xlrd.open_workbook(filename)
    sheet = wb.sheet_by_index(0)
    rows = sheet.nrows
    print(rows)
    cols = sheet.ncols
    print(cols)
    lst = []
    for row in range(rows):
        for col in range(cols):
            cell_data = sheet.cell_value(row, col)
            lst.append(cell_data)

    return lst
@pytest.mark.parametrize('id', get_data())
def test1(id):
    print(id)

# if __name__ == '__main__':
#     pytest.main(['-sv', 'test_excel.py'])




# if __name__ == '__main__':
#
#     filename=r'D:\pycharm-workspace\test_pytest\test_case\cases.xlsx'
#     s=readexecl(filename,'adminLogin')
#
#     print(s)