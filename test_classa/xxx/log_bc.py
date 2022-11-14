import pprint

import loguru
import openpyxl
from loguru import logger
import logging

# logger.add(sink='py54.log',encoding='utf-8',level='INFO',rotation='10 B')
# logging.basicConfig(level='INFO')   #设置日志等级，默认是warning级别
#
# logging.info('这是标准日志')
# logging.warning('警告日志')
# logging.error('错误日志')

# import logging
# #完整的logging流程
# #1，日志收集器logger
#
#
#
#
# logger = logging.getLogger('py56')  #获取日志
# logger.setLevel("INFO")   #设置日志等级
#
#
# #2  日志处理器，显示日记  本子
# handler=logging.StreamHandler()
# handler.setLevel("INFO")
#
# logger.addHandler(handler)  #
#
# #文件处理器，通过文件保存日志，另一个本子
#
# file_handler=logging.FileHandler(filename='l1.log',encoding='utf-8')
# file_handler.setLevel('INFO')
# logger.addHandler(file_handler)
#
#
# #格式化的
# fmt=logging.Formatter("%(asctime)s  |  %(levelname)s  | %(filename)s - %(message)s")
# handler.setFormatter(fmt)
# file_handler.setFormatter(fmt)
#

# from log_封装 import get_logger
# loggers=get_logger()
#
#
# loggers.info('这是标准日志')
# loggers.warning('警告日志')
# loggers.error('错误日志')


import logging
from openpyxl.worksheet.worksheet import Worksheet
import pprint



#
# def read_excel2(filename1,sheet_name1):
#
#     workbook=openpyxl.load_workbook(filename1)   #打开文件
#     # workbook.level("INFO")
#
#     sheet: Worksheet=workbook[sheet_name1]   #取选择sheet
#
#
#     data=list(sheet.values)  #取sheet值
#     title=data[0]  #取tiltle 第一行值
#     rows=data[1:]  #取后面的开始第一行之后所有
#
#
#     data=[dict(zip(title,row))  for row in rows]  #列表推导式，把两个列表压缩再一起，循环每一个元素，最后转换成字典
#
#     return data
#
# pprint.pprint(read_excel2("case33.xlsx","Sheet1"))

# def read_excel3(filename,sheetname):
#
#     workbook=openpyxl.load_workbook(filename)
#
#     sheet: Worksheet=workbook[sheetname]
#
#     data=list(sheet.values)
#     title=data[0]
#     rows=data[1:]
#
#     data=[]
#
#     for row in rows:
#         row_dict=dict(zip(title,row))
#         data.append(row_dict)
#
#     return data
#
# pprint.pprint(read_excel3("case33.xlsx","Sheet1"))
#

# for a in range(1,10):
#     for b in range(1,a+1):
#         print(f"{a}*{b}={a*b}",end="\t")
#
#     print()
#


# class da:
#      #类属性
#     champion="rng"
#     runner_up="skt"
#
#     def __init__(self,name,id,country):   #实例属性
#
#         self.name=name
#         self.id=id
#         self.country=country
#
#         print(f"获得本次第一名是{self.champion},上单选手{name}")
#         print(f"选手编号{id},他的国家是{country}")
#         print(f"获得第二名是{self.runner_up}")
#
#     def study(self,school,class_):   #实例方法
#         self.school=school
#         self.class_=class_
#
#         print(f"世界上最好的学校:{school}")
#         print(f"世界上最好的班级:{class_}")
#
#     def use(self):  #  调用其他实例方法  用self.
#
#         self.study("北京大学","8班")
#
#     @classmethod      #类方法
#
#     def actual(cls,work,salary):
#         cls.work=work
#         cls.salary=salary
#
#         print(f"工作:{work},薪资：{salary}")
#
#     @classmethod
#
#     def expected(cls):
#         # cls.study("红旗小学","6班")
#         cls.actual("点点点工作者","7293")   #只能调相同方法，类方法只能调用类
#
#     @staticmethod  #静态方法：和普通函数一致
#
#     def static(magic):
#         print(f"这是空的{magic}")
#
#
#
# a_da=da("李元浩","009","中国")   #实例对象，传参
#
# print(da)  #获取类
# print(da.champion)    #获取类属性
#
# print(a_da.name)  #获取实例属性
#
# a_da.study("清华大学","三年2班")
#
# a_da.use()
# a_da.actual("点工","7777")
# a_da.expected()
#
# da.static("魔术")


import openpyxl

workbook=openpyxl.load_workbook("case33.xlsx")
print(workbook)