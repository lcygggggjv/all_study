
# import openpyxl
# from openpyxl.worksheet.worksheet import Worksheet


# workbook=openpyxl.load_workbook("case33.xlsx")   #打开文件
# print(workbook)

#选择表单，表格sheet1
# sheet:Worksheet=workbook.get_sheet_by_name("Sheet1")
# sheet:Worksheet=workbook['Sheet1']  #列表取值
# print(sheet)

#获取某个单元格   cell单元格，row1行，一列
# cell= sheet.cell(row=1,column=1).value  #一行一列的值
# print(cell)

#获取所有用例数据
# rows=sheet.values
# for row in rows:
#     print(row)  #得到元组

# print(list(rows)[1:])  #打印数据转换列表,第一行过滤，切片

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

# def read_excel(file,sheet_name='Sheet1'):
#     workbook=openpyxl.load_workbook(file)  #不同文件名做参数，也可做默认参数
#
#     sheet:Worksheet=workbook[sheet_name]  #表格有不同名称
#
#     rows=list(sheet.values)[1:]
#
#     return rows  #取用例值
#
# print(read_excel("case33.xlsx"))  #要在控制台看，需要打印


# def read_excel_dict(file,sheet_name='Sheet1'):
#
#     workbook=openpyxl.load_workbook(file)
#
#     sheet:Worksheet=workbook[sheet_name]
#
#     data=list(sheet.values)
#     title=data[0]  #取第一行
#     rows=data[1:]  #从1开始取
#
#
#     data=[]  #定义一个空列表
#     for row in rows:
#         row_dict=dict(zip(title,row))  #zip把俩个列表压缩，组到一起，外面在用dict转换成字典
#         data.append(row_dict)  #把字典增加到空列表里
#
#     return data  #返回数据
#
#
# print(read_excel_dict('case33.xlsx'))  #打印


def read_dict(file,sheet_name='Sheet1'):   #形参file 默认sheet—nema

    workbook=openpyxl.load_workbook(file)  #先打开文件
    sheet:Worksheet=workbook[sheet_name]   #选择sheet页

    data=list(sheet.values)   #取sheet页里面值，转换成列表
    title=data[0]  #取第一行
    rows=data[1:]   #取从一行开始，到最后的

    data=[dict(zip(title,row))  for row in rows]  #列表推导式，循环每一个元素，用zip合并title，row，转换成字典，到新列表

    return data  #返回用例值

print(read_dict(r'D:\pycharm-workspace\test-框架搭建\datas\case3.xlsx'))   #调用函数，传参文件名，控制台print打印